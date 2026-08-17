import sys
import os
import logging
import time
import warnings
import numpy as np
import pandas as pd
import yfinance as yf
from datetime import datetime, timezone

warnings.filterwarnings("ignore")


# CONFIG
HEADLINES_LOG = "/Users/vinodgeorge/Desktop/headlines_log.csv"
TICKERS        = ["^GSPC", "AAPL", "MSFT", "NVDA", "AMD", "^VIX"]
OUTPUT_FILE    = "Sentiment_Report.xlsx"
CHARTS_FILE    = OUTPUT_FILE.replace(".xlsx", "_charts.xlsx")

# FinBERT model — downloads ~400MB on first run, then ca.ched locally
FINBERT_MODEL  = "ProsusAI/finbert"

# CSV log — headlines accumulate here across runs
# Keep this file alongside the script; don't delete it between runs
HEADLINES_LOG  = "headlines_log.csv"


# LOGGING
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)s  %(message)s",
    datefmt="%H:%M:%S",)
log = logging.getLogger(__name__)


# 1. NEWS INGESTION
log.info("Pulling news headlines ...")

def fetch_news(tickers: list[str]) -> pd.DataFrame:
    """
    Pull news headlines from yfinance for each ticker.
    Returns a DataFrame with columns:
        ticker, date, title, source, url
    """
    rows = []

    for ticker in tickers:
        # Skip index tickers — yfinance doesn't return news for ^GSPC, ^VIX etc.
        if ticker.startswith("^"):
            log.info("Skipping %s (index — no news feed)", ticker)
            continue

        log.info("Fetching news for %s ...", ticker)

        try:
            t    = yf.Ticker(ticker)
            news = t.get_news()

            if not news:
                log.warning("No news returned for %s", ticker)
                continue

            for article in news:
                # yfinance returns nested content dict in newer versions
                content = article.get("content", article)

                # Extract title
                title = (
                    content.get("title") or
                    article.get("title", "")
                ).strip()

                if not title:
                    continue

                # Extract timestamp — providerPublishTime is unix seconds
                ts = (
                    content.get("pubDate") or
                    article.get("providerPublishTime") or
                    article.get("pubDate"))
                if isinstance(ts, int):
                    date = datetime.fromtimestamp(ts, tz=timezone.utc).strftime("%Y-%m-%d")
                elif isinstance(ts, str):
                    date = ts[:10]   # trim to YYYY-MM-DD
                else:
                    date = datetime.now().strftime("%Y-%m-%d")

                # Extract source
                source = (
                    content.get("provider", {}).get("displayName") or
                    article.get("publisher", "Unknown"))

                url = (
                    content.get("canonicalUrl", {}).get("url") or
                    article.get("link", ""))

                rows.append({
                    "ticker":  ticker,
                    "date":    date,
                    "title":   title,
                    "source":  source,
                    "url":     url,})

            # Be polite to the API
            time.sleep(0.3)

        except Exception as e:
            log.warning("Failed to fetch news for %s: %s", ticker, e)
            continue

    if not rows:
        log.error("No headlines retrieved. Check your tickers and internet connection.")
        sys.exit(1)

    df = pd.DataFrame(rows)
    df["date"] = pd.to_datetime(df["date"])
    df = df.sort_values(["ticker", "date"]).reset_index(drop=True)

    log.info("Retrieved %d headlines across %d tickers", len(df), df["ticker"].nunique())
    return df


news_df = fetch_news(TICKERS)

# Quick inspection
print("\n── Raw headlines sample ──────────────────────────────────────────────")
print(news_df[["ticker", "date", "title", "source"]].head(10).to_string(index=False))
print(f"\nTotal headlines: {len(news_df)}")
print(f"Tickers with news: {news_df['ticker'].unique().tolist()}")
print(f"Date range: {news_df['date'].min().date()} → {news_df['date'].max().date()}")
print(f"Headlines per ticker:\n{news_df.groupby('ticker').size().to_string()}")


# 2. SENTIMENT SCORING (FinBERT)
log.info("Loading FinBERT model (downloads ~400MB on first run, cached after) ...")

from transformers import pipeline

try:
    finbert = pipeline(
        "text-classification",
        model=FINBERT_MODEL,
        tokenizer=FINBERT_MODEL,
        return_all_scores=True,   # all three probabilities, not just the winner
        truncation=True,
        max_length=512,           # FinBERT max context window
        device=-1,                # CPU — change to 0 if you have a GPU
)
    log.info("FinBERT loaded OK")
except Exception as e:
    log.error("Failed to load FinBERT: %s", e)
    log.error("Make sure you have run: pip install transformers torch")
    sys.exit(1)


def score_headlines(df: pd.DataFrame, batch_size: int = 16) -> pd.DataFrame:
    """
    Score each headline in df using FinBERT.

    FinBERT returns three probabilities per headline:
        positive, negative, neutral  (sum to 1.0)

    We store all three plus:
        label           — the winning class
        sentiment_score — composite score: positive - negative
                          range [-1, +1]; positive means bullish sentiment,
                          negative means bearish

    Batching avoids OOM on large headline sets and shows progress.
    """
    df = df.copy()
    headlines = df["title"].tolist()
    n = len(headlines)

    all_records = []

    for i in range(0, n, batch_size):
        batch = headlines[i : i + batch_size]
        log.info("Scoring headlines %d–%d of %d ...", i + 1, min(i + batch_size, n), n)

        try:
            raw = finbert(batch)
        except Exception as e:
            log.warning("Batch %d failed: %s — filling with neutral", i, e)
            raw = [[
                {"label": "positive", "score": 0.0},
                {"label": "negative", "score": 0.0},
                {"label": "neutral",  "score": 1.0},
            ]] * len(batch)

        for scores in raw:
            d     = {s["label"]: round(s["score"], 4) for s in scores}
            label = max(d, key=d.get)
            composite = round(d.get("positive", 0) - d.get("negative", 0), 4)
            all_records.append({
                "positive":        d.get("positive", 0),
                "negative":        d.get("negative", 0),
                "neutral":         d.get("neutral",  0),
                "label":           label,
                "sentiment_score": composite,})

    scores_df = pd.DataFrame(all_records)
    return pd.concat([df.reset_index(drop=True), scores_df], axis=1)


scored_df = score_headlines(news_df)

# Inspection 
print("\n── Scored headlines ─────────────────────────────────────────────────")
for _, row in scored_df[["ticker", "date", "title", "label", "sentiment_score"]].iterrows():
    title_short = row["title"][:55] + "..." if len(row["title"]) > 55 else row["title"]
    print(f"  {row['ticker']:5}  {str(row['date'].date()):12}  "
          f"{row['label']:8}  {row['sentiment_score']:+.3f}  {title_short}")

print(f"\nLabel distribution:")
print(scored_df["label"].value_counts().to_string())

print(f"\nMean sentiment score per ticker (+ = bullish, - = bearish):")
print(scored_df.groupby("ticker")["sentiment_score"].mean().round(3).to_string())


# 2b. CSV ACCUMULATION
# Append today's scored headlines to the running log.
# Deduplicates on URL so re-running the same day never double-counts.

def update_log(scored_df: pd.DataFrame, log_path: str) -> pd.DataFrame:
    """
    Append new scored headlines to the CSV log.
    Deduplicates on URL — safe to run multiple times per day.
    Returns the full accumulated log as a DataFrame.
    """
    if os.path.exists(log_path):
        existing = pd.read_csv(log_path, parse_dates=["date"])
        existing_count = len(existing)
        combined = pd.concat([existing, scored_df], ignore_index=True)
    else:
        existing_count = 0
        combined = scored_df.copy()
        log.info("Creating new headlines log: %s", log_path)

    # Deduplicate — prefer keeping rows with a real URL
    combined = combined.drop_duplicates(subset=["url"], keep="first")

    # Also deduplicate on ticker + date + title in case URL was empty
    combined = combined.drop_duplicates(subset=["ticker", "date", "title"], keep="first")

    combined = combined.sort_values(["ticker", "date"]).reset_index(drop=True)
    combined.to_csv(log_path, index=False)

    new_count = len(combined) - existing_count
    log.info("Log updated — %d total headlines (%d new this run)", len(combined), new_count)
    return combined


log_df = update_log(scored_df, HEADLINES_LOG)

print(f"\n── Accumulated log summary ───────────────────────────────────────────")
print(f"Total headlines in log : {len(log_df)}")
print(f"Date range             : {log_df['date'].min().date()} → {log_df['date'].max().date()}")
print(f"Headlines per ticker:")
print(log_df.groupby("ticker").size().to_string())
print(f"\nMonths covered per ticker:")
log_df["month"] = pd.to_datetime(log_df["date"]).dt.to_period("M")
print(log_df.groupby("ticker")["month"].nunique().to_string())



# 3. AGGREGATION & RETURN CORRELATION
log.info("Aggregating sentiment and computing return correlations ...")

# 3a. Load return data from equity analyser 
# Pulls the cleaned return series from your existing ticker_dfs_clean output.
# Run the equity analyser first so the data is available, or point
# EQUITY_FILE at an existing Financial_Data_Report.xlsx.
#
# If you want to run the sentiment analyser standalone, set EQUITY_FILE = None
# and we fall back to fetching returns directly from yfinance.

EQUITY_FILE = None   # set to "Financial_Data_Report.xlsx" if you want to read from it

def get_returns(tickers: list[str], start: str = "2024-01-01",
                end: str = "2026-05-31") -> pd.DataFrame:
    """
    Fetch monthly returns for each ticker directly from yfinance.
    Returns a long-format DataFrame: ticker, date (month start), return.
    """
    rows = []
    for ticker in tickers:
        try:
            df = yf.download(
                ticker, start=start, end=end,
                interval="1mo", auto_adjust=False,
                progress=False,)
            if df.empty:
                continue
            close = df["Close"].squeeze()
            ret   = close.pct_change().dropna()
            for date, r in ret.items():
                rows.append({
                    "ticker": ticker,
                    "month":  pd.Timestamp(date).to_period("M"),
                    "return": round(float(r), 6),})
        except Exception as e:
            log.warning("Could not fetch returns for %s: %s", ticker, e)

    return pd.DataFrame(rows)


# Only fetch returns for equity tickers (skip indices — no news for them anyway)
equity_tickers = [t for t in TICKERS if not t.startswith("^")]

log.info("Fetching monthly returns for correlation analysis ...")
returns_df = get_returns(equity_tickers)

if returns_df.empty:
    log.error("No return data retrieved — cannot compute correlations.")
else:
    log.info("Retrieved %d monthly return observations", len(returns_df))

# 3b. Monthly sentiment aggregation 
log_df["month"] = pd.to_datetime(log_df["date"]).dt.to_period("M")

monthly_sentiment = log_df.groupby(["ticker", "month"]).agg(
    mean_sentiment  = ("sentiment_score", "mean"),
    positive_pct    = ("label", lambda x: round((x == "positive").mean() * 100, 1)),
    negative_pct    = ("label", lambda x: round((x == "negative").mean() * 100, 1)),
    neutral_pct     = ("label", lambda x: round((x == "neutral").mean()  * 100, 1)),
    headline_count  = ("sentiment_score", "count"),
).round(3).reset_index()

print("\n── Monthly aggregated sentiment ─────────────────────────────────────")
print(monthly_sentiment.to_string(index=False))

# 3c. Join sentiment to returns
if not returns_df.empty:
    # Ensure Period dtype matches before merging
    returns_df["month"] = returns_df["month"].astype("period[M]")
    monthly_sentiment["month"] = monthly_sentiment["month"].astype("period[M]")

    merged = monthly_sentiment.merge(
        returns_df, on=["ticker", "month"], how="inner")

    # Forward return: sentiment in month T predicting return in month T+1
    merged = merged.sort_values(["ticker", "month"])
    merged["fwd_return"] = merged.groupby("ticker")["return"].shift(-1)

    # 3d. Correlation table
    corr_rows = []
    for ticker in equity_tickers:
        t = merged[merged["ticker"] == ticker].dropna(
            subset=["mean_sentiment", "return", "fwd_return"])
        if len(t) < 3:
            log.warning("%s: not enough overlapping months for correlation (need 3+)", ticker)
            continue

        corr_rows.append({
            "Ticker":               ticker,
            "Contemp. Corr (T)":    round(t["mean_sentiment"].corr(t["return"]),     3),
            "Forward Corr (T+1)":   round(t["mean_sentiment"].corr(t["fwd_return"]), 3),
            "Avg Sentiment":        round(t["mean_sentiment"].mean(), 3),
            "Avg Headlines/Month":  round(t["headline_count"].mean(), 1),
            "Months Observed":      len(t),})

    corr_df = pd.DataFrame(corr_rows)

    print("\n── Correlation: sentiment vs returns ────────────────────────────────")
    print(corr_df.to_string(index=False))

    # Pooled across all tickers
    all_data = merged.dropna(subset=["mean_sentiment", "return", "fwd_return"])
    pooled_contemp = all_data["mean_sentiment"].corr(all_data["return"])
    pooled_fwd     = all_data["mean_sentiment"].corr(all_data["fwd_return"])

    print(f"\nPooled contemporaneous correlation : {pooled_contemp:.3f}")
    print(f"Pooled forward correlation (T+1)   : {pooled_fwd:.3f}")
    print(f"Total monthly observations          : {len(all_data)}")
    print("""
Note: with ~10 headlines per ticker from yfinance, each month may have
only 1-2 observations — treat correlations as indicative, not conclusive.
Accumulate more headlines over time for a robust result.
""")
else:
    log.warning("Skipping correlation — no return data available.")
    merged  = monthly_sentiment.copy()
    corr_df = pd.DataFrame()


# 4. EXCEL OUTPUT
log.info("Building Excel workbook ...")

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import xlsxwriter

wb = Workbook()

# Styles 
HDR_FILL   = PatternFill(fill_type="solid", start_color="4F81BD", end_color="4F81BD")
HDR_FONT   = Font(bold=True, color="FFFFFF")
RED_FILL   = PatternFill(fill_type="solid", start_color="FFC7CE", end_color="FFC7CE")
GRN_FILL   = PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE")
GREY_FILL  = PatternFill(fill_type="solid", start_color="F2F2F2", end_color="F2F2F2")
THIN       = Side(style="thin")
BORDER     = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def write_table(ws, df, start_row=1, index=True):
    """Write a DataFrame as a formatted table. Returns next free row."""
    headers = ([df.index.name or ""] if index else []) + list(df.columns)
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=c, value=h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER

    for r_off, (idx, row_data) in enumerate(df.iterrows(), start=1):
        r = start_row + r_off
        col_start = 1
        if index:
            cell = ws.cell(row=r, column=1,
                           value=str(idx) if not hasattr(idx, "strftime") else idx.strftime("%Y-%m-%d"))
            cell.border = BORDER
            col_start = 2
        for c, val in enumerate(row_data, start=col_start):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = BORDER
            if r % 2 == 0:
                cell.fill = GREY_FILL

    # Auto-size columns
    for col in ws.columns:
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 3

    return start_row + len(df) + 3


def style_sentiment_col(ws, col_letter, data_start_row, data_end_row):
    """Colour sentiment_score cells green (positive) or red (negative)."""
    for r in range(data_start_row, data_end_row + 1):
        cell = ws[f"{col_letter}{r}"]
        if isinstance(cell.value, (int, float)):
            cell.fill = GRN_FILL if cell.value > 0 else (RED_FILL if cell.value < 0 else GREY_FILL)
            cell.number_format = "+0.000;-0.000;0.000"


# Sheet 1: Raw Headlines
ws1 = wb.active
ws1.title = "Raw Headlines"
ws1.freeze_panes = "A2"

display_cols = ["ticker", "date", "title", "source",
                "label", "sentiment_score", "positive", "negative", "neutral"]
raw_display  = log_df[[c for c in display_cols if c in log_df.columns]].copy()
raw_display["date"] = raw_display["date"].dt.strftime("%Y-%m-%d")

# Header
headers = list(raw_display.columns)
for c, h in enumerate(headers, start=1):
    cell = ws1.cell(row=1, column=c, value=h)
    cell.font = HDR_FONT
    cell.fill = HDR_FILL
    cell.alignment = Alignment(horizontal="center")
    cell.border = BORDER

# Data rows
for r, (_, row_data) in enumerate(raw_display.iterrows(), start=2):
    for c, val in enumerate(row_data, start=1):
        cell = ws1.cell(row=r, column=c, value=val)
        cell.border = BORDER
        if r % 2 == 0:
            cell.fill = GREY_FILL

# Colour label column
label_col = headers.index("label") + 1 if "label" in headers else None
score_col  = headers.index("sentiment_score") + 1 if "sentiment_score" in headers else None

for r in range(2, len(raw_display) + 2):
    if label_col:
        lc = ws1.cell(row=r, column=label_col)
        if lc.value == "positive":
            lc.fill = GRN_FILL
        elif lc.value == "negative":
            lc.fill = RED_FILL
    if score_col:
        sc = ws1.cell(row=r, column=score_col)
        if isinstance(sc.value, (int, float)):
            sc.fill = GRN_FILL if sc.value > 0 else (RED_FILL if sc.value < 0 else GREY_FILL)
            sc.number_format = "+0.000;-0.000;0.000"

# Auto-size
for col in ws1.columns:
    max_len = max((len(str(cell.value)) for cell in col if cell.value), default=8)
    letter  = get_column_letter(col[0].column)
    ws1.column_dimensions[letter].width = min(max_len + 3, 60)  # cap title col width

ws1.sheet_properties.tabColor = "808080"


# Sheet 2: Monthly Sentiment
ws2 = wb.create_sheet("Monthly Sentiment")
ws2.freeze_panes = "A2"

monthly_display = monthly_sentiment.copy()
monthly_display["month"] = monthly_display["month"].astype(str)
monthly_display = monthly_display.reset_index(drop=True)

write_table(ws2, monthly_display.set_index("ticker"), start_row=1, index=True)

# Colour mean_sentiment column
cols2 = ["ticker"] + list(monthly_sentiment.columns[1:])
if "mean_sentiment" in monthly_display.columns:
    ms_col_idx = list(monthly_display.columns).index("mean_sentiment") + 2  # +1 for index, +1 for 1-based
    ms_col_letter = get_column_letter(ms_col_idx)
    style_sentiment_col(ws2, ms_col_letter, 2, len(monthly_display) + 1)

ws2.sheet_properties.tabColor = "4F81BD"


# Sheet 3: Correlation Table 
ws3 = wb.create_sheet("Correlations")
ws3.freeze_panes = "A2"

if not corr_df.empty:
    write_table(ws3, corr_df.set_index("Ticker"), start_row=1, index=True)

    # Add pooled stats below
    next_r = len(corr_df) + 5
    ws3.cell(row=next_r,     column=1).value = "Pooled contemporaneous correlation"
    ws3.cell(row=next_r,     column=2).value = round(pooled_contemp, 3)
    ws3.cell(row=next_r + 1, column=1).value = "Pooled forward correlation (T+1)"
    ws3.cell(row=next_r + 1, column=2).value = round(pooled_fwd, 3)
    ws3.cell(row=next_r + 2, column=1).value = "Total monthly observations"
    ws3.cell(row=next_r + 2, column=2).value = len(all_data)

    # Note on interpretation
    ws3.cell(row=next_r + 4, column=1).value = (
        "Note: contemporaneous correlation reflects news that already moved prices. "
        "Forward correlation (T+1) tests whether sentiment predicts next-period returns. "
        "Weak forward correlation is consistent with semi-strong market efficiency.")
    ws3.cell(row=next_r + 4, column=1).font = Font(italic=True, size=9)
else:
    ws3.cell(row=1, column=1).value = "Insufficient data for correlation analysis."
    ws3.cell(row=2, column=1).value = "Run the script daily to accumulate headlines across 3+ months per ticker."
    ws3.cell(row=3, column=1).value = f"Current months in log: 1 — need at least 3."
    for r in range(1, 4):
        ws3.cell(row=r, column=1).font = Font(italic=True)

ws3.column_dimensions["A"].width = 35
ws3.sheet_properties.tabColor = "7030A0"


# Sheet 4: Sentiment Log Summary 
ws4 = wb.create_sheet("Log Summary")

# Headlines per ticker per month
pivot = (
    log_df.assign(month=pd.to_datetime(log_df["date"]).dt.to_period("M").astype(str))
    .groupby(["ticker", "month"])
    .agg(
        headlines     = ("title", "count"),
        mean_score    = ("sentiment_score", lambda x: round(x.mean(), 3)),
        pct_positive  = ("label", lambda x: round((x == "positive").mean() * 100, 1)),
        pct_negative  = ("label", lambda x: round((x == "negative").mean() * 100, 1)),)
    .reset_index())

ws4.cell(row=1, column=1).value = f"Headlines log: {HEADLINES_LOG}"
ws4.cell(row=1, column=1).font  = Font(bold=True)
ws4.cell(row=2, column=1).value = f"Total headlines: {len(log_df)}"
ws4.cell(row=3, column=1).value = f"Date range: {log_df['date'].min().date()} → {log_df['date'].max().date()}"

write_table(ws4, pivot.set_index("ticker"), start_row=5, index=True)
ws4.sheet_properties.tabColor = "00B050"


# Save openpyxl workbook 
wb.save(OUTPUT_FILE)
log.info("Saved → %s", OUTPUT_FILE)


# Charts file (xlsxwriter) 
log.info("Building charts file ...")

equity_tickers_chart = [t for t in TICKERS if not t.startswith("^")]
dates_chart = sorted(log_df["date"].dt.to_period("M").unique())

xw  = xlsxwriter.Workbook(CHARTS_FILE)
xws = xw.add_worksheet("Sentiment Charts")

# Write monthly sentiment data for charting
xws.write(0, 0, "Month")
for c, ticker in enumerate(equity_tickers_chart, start=1):
    xws.write(0, c, ticker)

for r, month in enumerate(dates_chart):
    xws.write(r + 1, 0, str(month))
    for c, ticker in enumerate(equity_tickers_chart, start=1):
        match = monthly_sentiment[
            (monthly_sentiment["ticker"] == ticker) &
            (monthly_sentiment["month"] == month)]
        val = match["mean_sentiment"].values[0] if len(match) else None
        if val is not None:
            xws.write(r + 1, c, round(float(val), 3))

n_months  = len(dates_chart)
chart_col = len(equity_tickers_chart) + 2

# Sentiment trend chart
c1 = xw.add_chart({"type": "line"})
c1.set_title({"name": "Monthly Mean Sentiment Score by Ticker"})
c1.set_x_axis({"name": "Month"})
c1.set_y_axis({"name": "Sentiment Score (+1 = bullish, -1 = bearish)"})
c1.set_style(10)
c1.set_size({"width": 700, "height": 380})

for c, ticker in enumerate(equity_tickers_chart, start=1):
    c1.add_series({
        "name":       ["Sentiment Charts", 0, c],
        "categories": ["Sentiment Charts", 1, 0, n_months, 0],
        "values":     ["Sentiment Charts", 1, c, n_months, c],
        "line":       {"width": 1.5},})

xws.insert_chart(0, chart_col, c1)

# % Positive headlines chart
POS_START_ROW = n_months + 4
xws.write(POS_START_ROW, 0, "Month")
for c, ticker in enumerate(equity_tickers_chart, start=1):
    xws.write(POS_START_ROW, c, ticker)

for r, month in enumerate(dates_chart):
    xws.write(POS_START_ROW + 1 + r, 0, str(month))
    for c, ticker in enumerate(equity_tickers_chart, start=1):
        match = monthly_sentiment[
            (monthly_sentiment["ticker"] == ticker) &
            (monthly_sentiment["month"] == month)]
        val = match["positive_pct"].values[0] if len(match) else None
        if val is not None:
            xws.write(POS_START_ROW + 1 + r, c, round(float(val), 1))

c2 = xw.add_chart({"type": "line"})
c2.set_title({"name": "% Positive Headlines by Ticker"})
c2.set_x_axis({"name": "Month"})
c2.set_y_axis({"name": "% Positive"})
c2.set_style(10)
c2.set_size({"width": 700, "height": 380})

for c, ticker in enumerate(equity_tickers_chart, start=1):
    c2.add_series({
        "name":       ["Sentiment Charts", POS_START_ROW, c],
        "categories": ["Sentiment Charts", POS_START_ROW + 1, 0, POS_START_ROW + n_months, 0],
        "values":     ["Sentiment Charts", POS_START_ROW + 1, c, POS_START_ROW + n_months, c],
        "line":       {"width": 1.5},})

xws.insert_chart(POS_START_ROW, chart_col, c2)
xw.close()

log.info("Charts saved → %s", CHARTS_FILE)
log.info("Done.")
